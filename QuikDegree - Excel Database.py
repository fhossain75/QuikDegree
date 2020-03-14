#Quik Degree - Franklin College of Arts and Sciences
#Author: Faisal Hossain

#Currently, demo is set to work for UGA's CS major requirements

# Import libraries
import requests
import urllib.request
import time 
import openpyxl
from openpyxl import Workbook 
from bs4 import BeautifulSoup

# Create workbook
wb = Workbook()

# Grab the active worksheet
worksheet = wb.active

# Set the URL you want to webscrape from
url = 'https://osas.franklin.uga.edu/franklin-college-degree-requirements'

# Connect to the URL
response = requests.get(url)

# Parse HTML and save to BeautifulSoup object
soup = BeautifulSoup(response.text, "html.parser")

# College Degree Requirements
start = soup.find('article') #find first html <article>
start1 = start.find('ul') # find start of bulleted listed in article

degreeReq = {} #{College Requirement:[Link, Requirements....],}
colCounter = 0
for a in start1.find_all('a'):
    colCounter += 1
    degreeReq[a.string] = ['https://osas.franklin.uga.edu/' + a.get('href')]
    worksheet.cell(row=1, column=colCounter).value = a.string

del degreeReq['Foreign Language Requirement\xa0'] #ERROR-Handling 
del degreeReq['History Requirement'] #ERROR-Handling 

# Print Statements
for key,value in degreeReq.items():
    print (key)
    print (value)
    print()

# College Degree Requirements Sub
print('**College Degree Requirements Sub - START')

colCounter = 0
rowCounter = 1
for key, value in degreeReq.items():
    url = value[0]
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    start = soup.find('tbody') #find first html <tbody>
    print()
    print("[|||||] " + key)
    colCounter += 1
    rowCounter = 1
    for a in start.find_all('a'): #Course
        rowCounter += 1
        worksheet.cell(row=rowCounter, column=colCounter).value = a.string
        

wb.save('QuikDegree.xlsx') # Save workbook
